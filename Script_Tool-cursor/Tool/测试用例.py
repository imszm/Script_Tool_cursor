import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# 定义输出文件名
file_name = "Walking_Assist_Test_Cases_Detailed.xlsx"

# 初始化用例列表
test_cases = []


def add_case(module, sub_module, item, pre_condition, steps, expected, priority="P1"):
    """辅助函数：添加一条测试用例"""
    test_cases.append({
        "ID": f"TC_{len(test_cases) + 1:03d}",
        "模块": module,
        "子模块": sub_module,
        "测试项": item,
        "前置条件": pre_condition,
        "测试步骤": steps,
        "预期结果": expected,
        "优先级": priority
    })


# ==========================================
# 1. 基础配置与开关逻辑 (对应导图 Source 63 附近)
# ==========================================
# 逻辑：电机使能开关 x 工具配置开关 的组合矩阵
add_case("基础逻辑", "配置开关", "组合1: 电机关+工具关",
         "1. 电机使能开关：OFF\n2. 工具配置开关：OFF",
         "1. 推动设备。",
         "推行不生效（无助力，无阻尼，离合脱开）。", "P0")

add_case("基础逻辑", "配置开关", "组合2: 电机关+工具开",
         "1. 电机使能开关：OFF\n2. 工具配置开关：ON",
         "1. 推动设备。",
         "推行不生效（优先级检查：电机使能为总开关）。", "P1")

add_case("基础逻辑", "配置开关", "组合3: 电机开+工具关",
         "1. 电机使能开关：ON\n2. 工具配置开关：OFF",
         "1. 推动设备。",
         "推行生效（正常助力模式）。", "P0")

add_case("基础逻辑", "配置开关", "组合4: 电机开+工具开",
         "1. 电机使能开关：ON\n2. 工具配置开关：ON",
         "1. 推动设备。",
         "推行生效（或进入特定的工具调试模式，需确认具体定义）。", "P2")

# ==========================================
# 2. 平地行驶逻辑 (对应导图左侧及中部)
# ==========================================
# 逻辑：静止起步 vs 行进中
add_case("平地控制", "静止起步", "平地-正常加速度起步",
         "平地，静止状态",
         "1. 推动设备，使加速度 < 0.188m/s²。\n2. 速度达到 0.8m/s 左右。",
         "1. 电机介入助力。\n2. 助力平滑，无突兀感。", "P0")

add_case("平地控制", "静止起步", "平地-急加速起步(防抖/误触)",
         "平地，静止状态",
         "1. 猛推设备，使加速度 > 0.188m/s²（模拟碰撞或误操作）。",
         "1. 不应产生助力。\n2. 或进入阻尼模式防止飞车。", "P1")

add_case("平地控制", "行进中", "平地-正常推行保持",
         "平地，速度在 0.8-1.2m/s 之间",
         "1. 保持匀速推行。",
         "1. 维持助力状态。\n2. 手感轻便。", "P0")

add_case("平地控制", "行进中", "平地-超速保护",
         "平地，助力状态",
         "1. 加速推行，使速度 > 1.2m/s。",
         "1. 助力停止。\n2. 产生反向阻尼（制动感）。\n3. 将速度限制在安全范围内。", "P0")

add_case("平地控制", "停止", "平地-自然减速停止",
         "平地，行进中",
         "1. 停止施加推力（但不松手刹）。",
         "1. 设备自然减速。\n2. 速度降至 0 时保持静止。", "P1")

# ==========================================
# 3. 坡道逻辑 (对应导图右侧大面积分支)
# ==========================================
# 导图明确区分了 3度, 6度, 9度, 12度。
# 这是一个典型的等价类划分，但导图将其展开了，所以我们也展开。

slopes = [3, 6, 9, 12]
for slope in slopes:
    # 上坡测试
    add_case("坡道控制", f"{slope}度坡", f"{slope}度-上坡正常助力",
             f"{slope}度坡道，车头朝上",
             "1. 向上推行，加速度 < 0.188m/s²。\n2. 速度保持 < 1.2m/s。",
             f"1. 上坡助力生效，推行省力。\n2. 能够克服{slope}度重力分量。", "P1" if slope < 12 else "P2")

    add_case("坡道控制", f"{slope}度坡", f"{slope}度-上坡急加速",
             f"{slope}度坡道，车头朝上",
             "1. 向上猛推，加速度 > 0.188m/s²。",
             "1. 识别为急加速。\n2. 限制助力输出或短暂保护。", "P2")

    add_case("坡道控制", f"{slope}度坡", f"{slope}度-坡道驻车(防后溜)",
             f"{slope}度坡道，推行中",
             "1. 松开双手（脱手）。",
             "1. 电磁刹车（抱闸）立即锁死。\n2. 设备无后溜。", "P0")

    # 下坡测试
    add_case("坡道控制", f"{slope}度坡", f"{slope}度-下坡匀速控制",
             f"{slope}度坡道，车头朝下",
             "1. 向下推行。",
             "1. 产生反向阻尼。\n2. 即使不拉住设备，速度也不应超过 1.2m/s。", "P1")

# ==========================================
# 4. 特殊场景与异常 (对应导图 Source 43, 53)
# ==========================================
add_case("异常处理", "单侧电机", "单侧加速度异常",
         "模拟单轮卡死或悬空",
         "1. 仅推动一侧扶手，造成左右加速度差值过大。",
         "1. 系统识别为异常操作。\n2. 不输出助力，防止原地快速打转。", "P1")

add_case("特殊场景", "小空间", "小空间原地转向",
         "狭窄空间（如电梯内）",
         "1. 操作设备原地旋转（左右轮反向运动）。",
         "1. 差速算法生效。\n2. 转向灵活，无机械卡滞感。", "P2")

add_case("特殊场景", "小空间", "小空间快速微调",
         "狭窄空间",
         "1. 快速、小幅度地前后调整位置。",
         "1. 响应灵敏。\n2. 不会出现助力延迟导致的“甚至冲出去”现象。", "P2")

# ==========================================
# 5. 压力与耐久性 (对应导图 Source 161-165)
# ==========================================
add_case("可靠性", "耐久测试", "连续推行1小时",
         "满负载（假人），平路或转鼓台",
         "1. 保持正常速度推行 1 小时。",
         "1. 电机温度在规格内。\n2. 无性能衰减。", "P2")

add_case("可靠性", "寿命测试", "启停/插拔100次",
         "静止状态",
         "1. 连续进行开关机操作 100 次。\n2. 或连续插拔手柄线束 100 次。",
         "1. 系统每次均能正常启动。\n2. 无死机、无报错。", "P2")

# ==========================================
# 生成 Excel 文件
# ==========================================
df = pd.DataFrame(test_cases)

try:
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='详细测试用例')

        workbook = writer.book
        worksheet = writer.sheets['详细测试用例']

        # 样式定义
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 深蓝色表头
        header_font = Font(bold=True, name="微软雅黑", color="FFFFFF")  # 白色字体
        body_font = Font(name="微软雅黑")
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # 列宽设置
        worksheet.column_dimensions['A'].width = 10  # ID
        worksheet.column_dimensions['B'].width = 12  # 模块
        worksheet.column_dimensions['C'].width = 15  # 子模块
        worksheet.column_dimensions['D'].width = 25  # 测试项
        worksheet.column_dimensions['E'].width = 30  # 前置条件
        worksheet.column_dimensions['F'].width = 45  # 测试步骤
        worksheet.column_dimensions['G'].width = 45  # 预期结果
        worksheet.column_dimensions['H'].width = 8  # 优先级

        # 应用样式
        for row in worksheet.iter_rows(min_row=1, max_row=len(test_cases) + 1):
            for cell in row:
                cell.alignment = alignment
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.font = body_font

    print(f"成功生成文件: {os.path.abspath(file_name)}")
    print(f"共生成 {len(test_cases)} 条测试用例，涵盖了 3/6/9/12 度坡度及开关组合逻辑。")

except Exception as e:
    print(f"Error: {e}")